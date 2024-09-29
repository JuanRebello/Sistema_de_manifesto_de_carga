#
import os
import pandas as pd
from tkinter import *
from datetime import datetime
from openpyxl import load_workbook, Workbook
import win32com.client as win32

class Application:
    def __init__(self):
        self.root = Tk()
        self.tela()
        self.frames_da_tela()
        self.criando_botoes()
        self.nome_do_campo_escrevente()
        self.numeros_de_copias()
        self.menu_aeronaves()
        self.menu_pcandestino()
        self.workbook = None

        # Caminhos dos arquivos
        self.arquivo_referencia = r"C:\Users\rebello\Desktop\scrapy siloms\dados_carga.xlsx" # aqui voce define o local do arquivo gerado pelo o scrapy feito
        self.arquivo_manifesto = r"C:\Users\rebello\Desktop\scrapy siloms\MANIFESTO MANUAL(1).xlsx" # aqui o modelo de arquivo que deseja inserir os dados

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.mainloop()

    def tela(self):
        self.root.title("Manifesto Manual")
        self.root.configure(background='#000000')
        self.root.geometry('400x300')
        self.root.resizable(False, False)

    def frames_da_tela(self):
        self.frame_1 = Frame(self.root, bd=4, bg='#000000', highlightbackground='#004C99', highlightthickness=3)
        self.frame_1.place(relx=0.02, rely=0.02, relwidth=0.96, relheight=0.96)

    def criando_botoes(self):
        self.bt_imprimir_manifesto = Button(self.frame_1, text='Imprimir Manifesto', bd=4, bg='#0066CC', fg='white', activebackground='#0059B3', activeforeground='white', command=self.imprimir_manifesto)
        self.bt_imprimir_manifesto.place(relx=0.4, rely=0.5, relwidth=0.3, relheight=0.08)
        self.bt_adicionar_volume = Button(self.frame_1, text='Adicionar Volume', bd=4, bg='#0066CC', fg='white', activebackground='#0059B3', activeforeground='white', command=self.abrir_janela_adicionar_volume)
        self.bt_adicionar_volume.place(relx=0.08, rely=0.5, relwidth=0.3, relheight=0.08)
        

    def nome_do_campo_escrevente(self):
        self.lb_codigo_1 = Label(self.frame_1, text='Aeronave', bg='#000000', fg='white')
        self.lb_codigo_1.place(relx=0.03, rely=0.03)
        self.lb_codigo_2 = Label(self.frame_1, text="Pcan Destino", bg='#000000', fg='white')
        self.lb_codigo_2.place(relx=0.45, rely=0.03)
        


    def numeros_de_copias(self):
        self.lb_codigo_a = Label(self.frame_1, text='Nº Cópias', bg='#000000', fg='white')
        self.lb_codigo_a.place(relx=0.8, rely=0.3)
        self.codigo_spinbox = Spinbox(self.frame_1, from_=1, to=10, bg='#e6f0ff', fg='black')
        self.codigo_spinbox.place(relx=0.81, rely=0.38, relwidth=0.1, relheight=0.05)

    def carregar_workbook(self):
        if not os.path.exists(self.arquivo_manifesto):
            print("Arquivo de manifesto não encontrado.")
            return
        self.workbook = load_workbook(self.arquivo_manifesto)
    
    def menu_aeronaves(self):
        self.opcoes_aeronaves = [" ", "C-95", "C-97", "C-98", "C-99", "C-105", "C-130", 'KC-30', 'KC-390', "Outros", "Terrestre"]
        self.opcao_aeronave = StringVar()
        self.opcao_aeronave.set(self.opcoes_aeronaves[0])
        
        self.menu_aeronaves = OptionMenu(self.frame_1, self.opcao_aeronave, *self.opcoes_aeronaves, command=self.atualizar_celula_a7)
        self.menu_aeronaves.place(relx=0.04, rely=0.11, relwidth=0.15, relheight=0.05)
        self.menu_aeronaves.config(bg='#e6f0ff', fg='black')
        self.menu_aeronaves['state'] = 'disabled'  # Desativa o menu inicialmente

    def atualizar_celula_a7(self, _):
        self.carregar_workbook()  # Carregar o workbook antes de atualizar
        opcao_selecionada = self.opcao_aeronave.get()
        if not hasattr(self, 'workbook') or self.workbook is None:
            print("Workbook não carregado.")
            return

        for sheet in self.workbook.sheetnames[1:]:
            ws = self.workbook[sheet]
            ws['A4'] = opcao_selecionada

        self.workbook.save(self.arquivo_manifesto)
        print("Célula A7 atualizada em todos os sheets.")
        

    def menu_pcandestino(self):
        self.opcoes = [" ", "CTLA", "ECAN-BQ", "ECAN-FZ", "ECAN-GW", "PCAN-BR", "PCAN-AF", "PCAN-AK", "PCAN-AN", "PCAN-BE", "PCAN-BR", "PCAN-BV", "PCAN-CABE", "PCAN-CABW", "PCAN-CC", "PCAN-CG", "PCAN-CO", "PCAN-CT", "PCAN-FL", "PCAN-FN", "PCAN-GR", "PCAN-LS", "PCAN-MN", "PCAN-NT", "PCAN-PV", "PCAN-RF", "PCAN-SC", "PCAN-SJ", "PCAN-SM", "PCAN-ST", "PCAN-SV", "PCAN-YS", "TCTL-GL", "TTL-EI", "TTL-TT", "TTL-UA"]
        self.opcao = StringVar()
        self.opcao.set(self.opcoes[0])

        self.menu_pcandestino = OptionMenu(self.frame_1, self.opcao, *self.opcoes)
        self.menu_pcandestino.place(relx=0.46, rely=0.11 ,relwidth=0.15, relheight=0.05)
        self.menu_pcandestino.config(bg='#e6f0ff', fg='black')

    def abrir_janela_adicionar_volume(self):
        opcao_destino = self.opcao.get()
        if opcao_destino == " ":
            self.notificar_sem_volumes()
            return

        self.nova_janela = Toplevel(self.root)
        self.nova_janela.title(f"Adicionar Volume - {opcao_destino}")
        self.nova_janela.configure(background='#004C99')
        self.nova_janela.geometry('600x400')

        self.canvas = Canvas(self.nova_janela, bg='#004C99')
        self.canvas.pack(side=LEFT, fill=BOTH, expand=True)

        self.scrollbar = Scrollbar(self.nova_janela, orient=VERTICAL, command=self.canvas.yview)
        self.scrollbar.pack(side=RIGHT, fill=Y)

        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind('<Configure>', lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        self.frame_dados = Frame(self.canvas, bg='#003366')
        self.canvas.create_window((0, 0), window=self.frame_dados, anchor="nw")

        self.bt_salvar_selecionados = Button(self.nova_janela, text='Salvar Selecionados', bd=4, bg='#0066CC', fg='white', command=self.salvar_e_fechar )
        self.bt_salvar_selecionados.place(relx=0.65, rely=0.8, relwidth=0.3, relheight=0.05)
        self.menu_aeronaves['state'] = 'normal'
        self.carregar_volumes_referencia()

    def salvar_e_fechar(self):
        self.salvar_selecionados()  # Executa o método para salvar as informações
        self.nova_janela.destroy()  # Fecha a janela após salvar    

    def carregar_volumes_referencia(self):
        opcao_destino = self.opcao.get()
        if not os.path.exists(self.arquivo_referencia):
            print("Arquivo de referência não encontrado.")
            return

        df = pd.read_excel(self.arquivo_referencia, sheet_name=opcao_destino)
        if df.empty:
            self.notificar_sem_volumes()
            return

        self.checkbuttons_vars = []
        for index, row in df.iterrows():
            var = IntVar()
            cb = Checkbutton(self.frame_dados, text=row['Volume'], variable=var)
            cb.grid(row=index, column=0, sticky='w')
            self.checkbuttons_vars.append(var)

    def salvar_selecionados(self):
        opcao_selecionada = self.opcao.get()  # Obtém a opção selecionada do menu

        if not os.path.exists(self.arquivo_referencia):
            print("Arquivo de referência não encontrado.")
            return

        if not os.path.exists(self.arquivo_manifesto):
            print("Arquivo de manifesto não encontrado.")
            return

        try:
            # Carregar dados do arquivo de referência
            wb_referencia = load_workbook(self.arquivo_referencia)

            if opcao_selecionada in wb_referencia.sheetnames:
                sheet_referencia = wb_referencia[opcao_selecionada]

                # Carregar dados do arquivo de manifesto
                wb_manifesto = load_workbook(self.arquivo_manifesto)
                original_sheet = wb_manifesto.active
                original_sheet_title = original_sheet.title

                # Variável para controlar a próxima linha disponível no manifesto
                next_row = 6

                # Contador para os sheets
                sheet_count = 1

                # Adicionar um novo sheet para garantir que sempre haverá uma planilha disponível
                new_sheet = wb_manifesto.copy_worksheet(original_sheet)
                new_sheet.title = f"{original_sheet_title}_{sheet_count}"
                sheet_manifesto = new_sheet

                # Processar linhas selecionadas
                for idx, var in enumerate(self.checkbuttons_vars):
                    if var.get() == 1:
                        # Verificar se ultrapassou o limite de 10 linhas
                        if next_row > 15:
                            # Adicionar um novo sheet com o layout original
                            sheet_count += 1
                            new_sheet = wb_manifesto.copy_worksheet(original_sheet)
                            new_sheet.title = f"{original_sheet_title}_{sheet_count}"
                            sheet_manifesto = new_sheet
                            next_row = 6  # Reiniciar o contador de linha

                        # Definir a linha de origem no arquivo de referência
                        source_row = 2 + idx

                        # Ler os dados das células da linha correspondente do arquivo de referência
                        dados = [
                            sheet_referencia[f'A{source_row}'].value,
                            sheet_referencia[f'B{source_row}'].value,
                            sheet_referencia[f'C{source_row}'].value,
                            sheet_referencia[f'D{source_row}'].value,
                        ]

                        # Tratar o valor da célula 'G' para garantir que seja float
                        valor_g = str(dados[3]).replace(",", ".").strip()
                        try:
                            dados[3] = float(valor_g)  # Converte o valor para float
                        except ValueError:
                            dados[3] = None  # Define como None se houver um erro na conversão

                        # Inserir os dados no arquivo de manifesto nas células correspondentes
                        sheet_manifesto[f'C{next_row}'] = dados[0]  # Valor da célula A na linha correspondente
                        sheet_manifesto[f'B{next_row}'] = dados[2]  # Valor da célula C na linha correspondente
                        sheet_manifesto[f'A{next_row}'] = dados[1]  # Valor da célula B na linha correspondente
                        sheet_manifesto[f'G{next_row}'] = dados[3]  # Valor da célula D na linha correspondente

                        # Incrementar a linha para a próxima entrada
                        next_row += 1

                # Salvar os dados no arquivo de manifesto
                wb_manifesto.save(self.arquivo_manifesto)
                print("Dados adicionados ao manifesto manual.")

            else:
                print(f"Não há informações para a opção selecionada: {opcao_selecionada}")

        except Exception as e:
            print(f"Erro ao salvar os dados no arquivo de manifesto: {e}")



    def carregar_volumes_manifesto(self):
        if not os.path.exists(self.arquivo_manifesto):
            print("Arquivo de manifesto não encontrado.")
            return

        wb = load_workbook(self.arquivo_manifesto)
        ws = wb.active

        self.checkbuttons_vars = []
        for index, row in enumerate(ws.iter_rows(values_only=True)):
            var = IntVar()
            cb = Checkbutton(self.frame_dados, text=row[0], variable=var)
            cb.grid(row=index, column=0, sticky='w')
            self.checkbuttons_vars.append(var)

    def imprimir_manifesto(self):
        try:
            numero_de_copias = int(self.codigo_spinbox.get())  # Obter o número de cópias do Spinbox
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False  # Não mostrar o Excel

            # Abrir o arquivo manifesto
            workbook = excel.Workbooks.Open(os.path.abspath(self.arquivo_manifesto))

            # Iterar sobre os sheets
            for sheet in workbook.Sheets:
                if sheet.Name != 'Planilha1':  # Ignorar o sheet com nome 'Planilha1'
                    sheet.PageSetup.Orientation = win32.constants.xlLandscape
                    for _ in range(numero_de_copias):
                        sheet.PrintOut()  # Imprimir a planilha atual
                        

            # Fechar o arquivo sem salvar alterações
            workbook.Close(SaveChanges=False)
            excel.Application.Quit()
            print("Sucesso", f"{numero_de_copias} cópias de cada sheet (exceto 'Planilha1') enviadas para impressão.")

        except Exception as e:
            print("Erro", f"Não foi possível imprimir o arquivo: {e}")





    def notificar_sem_volumes(self):
        self.nova_janela = Toplevel(self.root)
        self.nova_janela.title("Aviso")
        self.nova_janela.geometry("300x100")
        self.nova_janela.configure(background='#004C99')

        label = Label(self.nova_janela, text="Não há volumes disponíveis para a seleção.", bg='#004C99', fg='white')
        label.pack(pady=20)

        btn_ok = Button(self.nova_janela, text="OK", command=self.nova_janela.destroy)
        btn_ok.pack()

    def limpar_conteudo_manifesto(self):
        try:
            # Carregar o arquivo manifesto
            wb_manifesto = load_workbook(self.arquivo_manifesto)

            # Identificar o nome da planilha original
            sheet_original_name = wb_manifesto.sheetnames[0]

            # Iterar por todas as planilhas no arquivo
            for sheet_name in wb_manifesto.sheetnames:
                sheet = wb_manifesto[sheet_name]

                # Limpar o conteúdo das linhas 9 até 18
                for row in sheet.iter_rows(min_row=6, max_row=15):
                    for cell in row:
                        # Verifica se a célula não faz parte de uma célula mesclada
                        if not any(cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges):
                            cell.value = None

                # Remover sheets que não sejam o original
                if sheet_name != sheet_original_name:
                    wb_manifesto.remove(sheet)

            # Salvar o arquivo manifesto após a limpeza
            wb_manifesto.save(self.arquivo_manifesto)
            print("Conteúdo do arquivo manifesto limpo com sucesso. Sheets adicionais removidos.")

        except Exception as e:
            print(f"Erro ao limpar o conteúdo do arquivo_manifesto: {e}")

    def on_closing(self):
        # Limpar o conteúdo do arquivo_manifesto antes de fechar o programa
        self.limpar_conteudo_manifesto()
        self.root.destroy()



if __name__ == "__main__":
    app = Application()
